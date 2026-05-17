import json
import urllib.request
import urllib.error
import datetime
import openpyxl

EXCEL_PATH = r"C:\Projects\factory-map\uploads\IPC_data (1).xlsx"
API_BASE = "http://localhost:4000/api"

def api_post(path, payload, token=None):
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(f"{API_BASE}{path}", data=data, method="POST")
    req.add_header("Content-Type", "application/json")
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode("utf-8"))

def str_val(cell):
    """Return stripped string or empty string for None/empty cells."""
    if cell is None:
        return ""
    if isinstance(cell, (datetime.datetime, datetime.date)):
        return cell.strftime("%Y-%m-%d")
    return str(cell).strip()

def login():
    resp = api_post("/auth/login", {"username": "admin", "password": "Admin@1234"})
    return resp.get("token") or resp.get("data", {}).get("token")

def row_to_asset_cummins_motor(row, sheet_name):
    """IPC_Cummins and IPC_Motor share the same column layout."""
    object_id      = str_val(row[0])
    hw_asset_id    = str_val(row[1])
    os_version     = str_val(row[2])
    manufacturer   = str_val(row[3])
    model          = str_val(row[4])
    serial_number  = str_val(row[5])
    vlan           = str_val(row[6])
    remote_access  = str_val(row[7])
    backup_tool    = str_val(row[8])
    hw_specs       = str_val(row[9])
    # col 10 = SW list, skip
    environment    = str_val(row[11]) if len(row) > 11 else ""
    serial_object  = str_val(row[12]) if len(row) > 12 else ""
    notes          = str_val(row[13]) if len(row) > 13 else ""
    fortiedr       = str_val(row[14]) if len(row) > 14 else ""
    winupdate      = str_val(row[15]) if len(row) > 15 else ""

    # Skip completely empty rows
    if not object_id and not hw_asset_id and not serial_object:
        return None

    display_name = object_id or hw_asset_id or serial_object

    return {
        "basic_info": {
            "display_name": display_name,
            "asset_type": "IPC",
            "status": "active",
            "source_sheet": sheet_name,
        },
        "hierarchy": {},
        "location": {"coordinates": {"x": 0, "y": 0}},
        "itsm": {
            "hardware_asset_id": hw_asset_id,
            "is_managed": False,
            "sync_status": "never",
        },
        "network": {
            "ip_address": "",
            "vlan": vlan,
        },
        "technical_specs": {
            "os_version": os_version,
            "manufacturer": manufacturer,
            "model": model,
            "serial_number": serial_number,
            "cpu": hw_specs,
        },
        "custom_fields": {
            "object_id": object_id,
            "serial_object": serial_object,
            "environment": environment,
            "notes": notes,
            "remote_access_tool": remote_access,
            "backup_tool": backup_tool,
            "fortiedr_active": fortiedr,
            "winupdate_date": winupdate,
        },
    }

def row_to_asset_rotor(row):
    """IPC_Rotor: col 0 = display_name (Gép név), col 12 = serial_object."""
    display_name   = str_val(row[0])
    hw_asset_id    = str_val(row[1]) if len(row) > 1 else ""
    os_version     = str_val(row[2]) if len(row) > 2 else ""
    manufacturer   = str_val(row[3]) if len(row) > 3 else ""
    model          = str_val(row[4]) if len(row) > 4 else ""
    serial_number  = str_val(row[5]) if len(row) > 5 else ""
    vlan           = str_val(row[6]) if len(row) > 6 else ""
    remote_access  = str_val(row[7]) if len(row) > 7 else ""
    backup_tool    = str_val(row[8]) if len(row) > 8 else ""
    hw_specs       = str_val(row[9]) if len(row) > 9 else ""
    # col 10 = SW, skip
    environment    = str_val(row[11]) if len(row) > 11 else ""
    serial_object  = str_val(row[12]) if len(row) > 12 else ""
    notes          = str_val(row[13]) if len(row) > 13 else ""
    fortiedr       = str_val(row[14]) if len(row) > 14 else ""
    winupdate      = str_val(row[15]) if len(row) > 15 else ""

    if not display_name and not serial_object and not hw_asset_id:
        return None

    return {
        "basic_info": {
            "display_name": display_name or serial_object or hw_asset_id,
            "asset_type": "IPC",
            "status": "active",
            "source_sheet": "IPC_Rotor",
        },
        "hierarchy": {},
        "location": {"coordinates": {"x": 0, "y": 0}},
        "itsm": {
            "hardware_asset_id": hw_asset_id,
            "is_managed": False,
            "sync_status": "never",
        },
        "network": {
            "ip_address": "",
            "vlan": vlan,
        },
        "technical_specs": {
            "os_version": os_version,
            "manufacturer": manufacturer,
            "model": model,
            "serial_number": serial_number,
            "cpu": hw_specs,
        },
        "custom_fields": {
            "object_id": "",
            "serial_object": serial_object,
            "environment": environment,
            "notes": notes,
            "remote_access_tool": remote_access,
            "backup_tool": backup_tool,
            "fortiedr_active": fortiedr,
            "winupdate_date": winupdate,
        },
    }

def read_sheet(ws, sheet_name):
    assets = []
    rows = list(ws.iter_rows(values_only=True))
    # Skip header row(s): first row with data starts at index 1
    for i, row in enumerate(rows[1:], start=2):
        try:
            if sheet_name == "IPC_Rotor":
                asset = row_to_asset_rotor(row)
            else:
                asset = row_to_asset_cummins_motor(row, sheet_name)
            if asset:
                assets.append(asset)
        except Exception as e:
            print(f"  [WARN] Row {i} skipped: {e}")
    return assets

def bulk_import(assets, token, batch_size=50):
    total = len(assets)
    success = 0
    failed = 0
    for start in range(0, total, batch_size):
        batch = assets[start:start + batch_size]
        try:
            resp = api_post("/assets/bulk", {"assets": batch}, token)
            results = resp.get("data", {}).get("results", [])
            for r in results:
                if r.get("success"):
                    success += 1
                else:
                    failed += 1
                    asset_info = r.get("asset", {})
                    name = asset_info.get("basic_info", {}).get("display_name", "?")
                    custom = asset_info.get("custom_fields", {})
                    print(f"  [FAIL] {r.get('error', 'unknown')} — name: {name!r}, winupdate: {custom.get('winupdate_date','')!r}")
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8")
            print(f"  [HTTP ERROR {e.code}] {body[:400]}")
            failed += len(batch)
        except Exception as e:
            print(f"  [ERROR] {e}")
            failed += len(batch)
    return success, failed

def main():
    print("Logging in...")
    token = login()
    if not token:
        print("ERROR: Could not obtain auth token")
        return

    print(f"Token obtained. Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    ipc_sheets = {
        "IPC_Cummins": "IPC_Cummins",
        "IPC_Motor": "IPC_Motor",
        "IPC_Rotor": "IPC_Rotor",
    }

    all_assets = []
    for sheet_name in ipc_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        assets = read_sheet(ws, sheet_name)
        print(f"  {sheet_name}: {len(assets)} assets parsed")
        all_assets.extend(assets)

    wb.close()
    print(f"\nTotal assets to import: {len(all_assets)}")
    print("Importing...")

    success, failed = bulk_import(all_assets, token)
    print(f"\nDone. Success: {success}, Failed: {failed}")

if __name__ == "__main__":
    main()
